import asyncio
from playwright.async_api import async_playwright
import openpyxl

# Hàm đọc dữ liệu từ file Excel
def read_test_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua hàng tiêu đề
        if row[0] is not None:
            data.append(row)
    return data

# Hàm ghi kết quả test vào file Excel
def write_test_results(file_path, results):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    for i, result in enumerate(results, start=2):  # Ghi bắt đầu từ hàng thứ 2
        sheet.cell(row=i, column=4, value=result)
    workbook.save(file_path)

async def test_signup_and_login(playwright, data):
    browser = await playwright.chromium.launch(headless=False)
    page = await browser.new_page()

    results = []

    for test_case in data:
        name, email, password = test_case

        # Test trang Sign Up
        await page.goto('https://phptravels.net/')
        await page.click("text=Signup")
        await page.fill("input[name='first_name']", name)
        await page.fill("input[name='email']", email)
        await page.fill("input[name='password']", password)
        await page.click("button[type='submit']")
        
        # Kiểm tra kết quả Sign Up
        if "Thank you for signing up" in await page.content():
            results.append("Sign Up Successful")
        else:
            results.append("Sign Up Failed")
        
        # Test trang Login
        await page.goto('https://phptravels.net/')
        await page.click("text=Login")
        await page.fill("input[name='email']", email)
        await page.fill("input[name='password']", password)
        await page.click("button[type='submit']")
        
        # Kiểm tra kết quả Login
        if "Welcome back" in await page.content():
            results[-1] += " | Login Successful"
        else:
            results[-1] += " | Login Failed"

    await browser.close()
    return results

async def main():
    file_path = 'test_data.xlsx'  # Đường dẫn tới file Excel chứa test data
    test_data = read_test_data(file_path)

    async with async_playwright() as playwright:
        results = await test_signup_and_login(playwright, test_data)

    write_test_results(file_path, results)
    print("Test completed. Results have been written to the Excel file.")

if __name__ == "__main__":
    asyncio.run(main())
